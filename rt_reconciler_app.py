import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os

class ReconcilerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("RT Reconciler")
        self.root.geometry("520x400")
        self.root.resizable(False, False)
        self.root.configure(bg='#f0f0f0')
        self.simple_file = tk.StringVar()
        self.rt_file = tk.StringVar()
        self.create_widgets()

    def create_widgets(self):
        title = tk.Label(self.root, text="RT vs Simple Reconciler", font=("Arial", 18, "bold"), bg='#f0f0f0')
        title.pack(pady=20)
        frame1 = tk.Frame(self.root, bg='#f0f0f0', pady=10)
        frame1.pack(fill="x", padx=20)
        tk.Label(frame1, text="Simple Workbook:", font=("Arial", 10), bg='#f0f0f0').pack(anchor="w")
        f1 = tk.Frame(frame1, bg='#f0f0f0')
        f1.pack(fill="x", pady=5)
        tk.Entry(f1, textvariable=self.simple_file, width=55, font=("Arial", 9)).pack(side="left", fill="x", expand=True)
        tk.Button(f1, text="Browse...", command=self.browse_simple, width=10).pack(side="right", padx=5)
        frame2 = tk.Frame(self.root, bg='#f0f0f0', pady=10)
        frame2.pack(fill="x", padx=20)
        tk.Label(frame2, text="RT Scan Export:", font=("Arial", 10), bg='#f0f0f0').pack(anchor="w")
        f2 = tk.Frame(frame2, bg='#f0f0f0')
        f2.pack(fill="x", pady=5)
        tk.Entry(f2, textvariable=self.rt_file, width=55, font=("Arial", 9)).pack(side="left", fill="x", expand=True)
        tk.Button(f2, text="Browse...", command=self.browse_rt, width=10).pack(side="right", padx=5)
        self.status_var = tk.StringVar(value="Select files and click Reconcile")
        tk.Label(self.root, textvariable=self.status_var, font=("Arial", 10), bg='#f0f0f0').pack(pady=15)
        self.btn = tk.Button(self.root, text="RECONCILE", font=("Arial", 14, "bold"), width=20, height=2, bg='#4472C4', fg='white', activebackground='#3461b3', activeforeground='white', cursor='hand2', command=self.start_reconcile)
        self.btn.pack(pady=20)

    def browse_simple(self):
        f = filedialog.askopenfilename(title="Select Simple Workbook", filetypes=[("Excel", "*.xlsx *.xls")])
        if f: self.simple_file.set(f)

    def browse_rt(self):
        f = filedialog.askopenfilename(title="Select RT Scan Export", filetypes=[("Excel", "*.xlsx *.xls")])
        if f: self.rt_file.set(f)

    def start_reconcile(self):
        if not self.simple_file.get() or not self.rt_file.get():
            messagebox.showerror("Error", "Please select both files")
            return
        self.btn.config(state='disabled')
        self.status_var.set("Processing...")
        self.root.update()
        try:
            output, stats = self.reconcile(self.simple_file.get(), self.rt_file.get())
            self.on_complete(output, stats)
        except Exception as e:
            import traceback
            self.on_error(str(e) + "\n\n" + traceback.format_exc())

    def on_complete(self, output_file, stats):
        self.btn.config(state='normal')
        self.status_var.set("Complete!")
        msg = "Reconciliation complete!\n\n"
        msg += "Total rows in: {}\n".format(stats['total_in'])
        msg += "IE Tire (not yet scanned): {}\n".format(stats['remaining'])
        msg += "Ready to Receive: {}\n".format(stats['ready'])
        msg += "Previously Received: {}\n".format(stats['prev_received'])
        msg += "Unmatched Scans: {}\n".format(stats['unmatched'])
        msg += "Total rows out: {}\n\n".format(stats['total_out'])
        msg += "Output: {}\n\nOpen now?".format(os.path.basename(output_file))
        if messagebox.askyesno("Success", msg):
            os.startfile(output_file)

    def on_error(self, msg):
        self.btn.config(state='normal')
        self.status_var.set("Error")
        messagebox.showerror("Error", msg)

    @staticmethod
    def clean_part(val):
        if pd.isna(val):
            return ''
        return str(val).strip().rstrip('[').upper()

    def reconcile(self, simple_file, rt_file):
        xl = pd.ExcelFile(simple_file)
        ie_df = pd.read_excel(xl, sheet_name='IE Tire')
        # Combine all detail tabs into one pool so no rows get lost
        all_detail = ie_df.copy()
        for name in ['Ready_to_Receive', 'Ready to Receive']:
            if name in xl.sheet_names:
                all_detail = pd.concat([all_detail, pd.read_excel(xl, sheet_name=name)], ignore_index=True)
                break
        for name in ['Previously Received', 'Previously_Received']:
            if name in xl.sheet_names:
                all_detail = pd.concat([all_detail, pd.read_excel(xl, sheet_name=name)], ignore_index=True)
                break
        # Load existing unmatched scans to carry forward
        existing_unmatched = pd.DataFrame()
        for name in ['Unmatched_Scans', 'Unmatched']:
            if name in xl.sheet_names:
                existing_unmatched = pd.read_excel(xl, sheet_name=name)
                break
        total_in = len(all_detail)
        all_detail['_clean_iet'] = all_detail['IET #'].apply(self.clean_part)
        if 'part_number' in all_detail.columns:
            all_detail['_clean_pn'] = all_detail['part_number'].apply(self.clean_part)
        else:
            all_detail['_clean_pn'] = ''
        # Load and aggregate RT scans
        rt_raw = pd.read_excel(rt_file, sheet_name=0)
        part_col = None
        qty_col = None
        for c in rt_raw.columns:
            cl = str(c).lower().strip()
            if cl == 'part':
                part_col = c
            elif 'qty' in cl or 'quantity' in cl:
                qty_col = c
        if part_col is None:
            raise ValueError("Cannot find Part column in RT file. Columns: {}".format(rt_raw.columns.tolist()))
        if qty_col is None:
            qty_col = '_qty'
            rt_raw[qty_col] = 1
        else:
            rt_raw[qty_col] = pd.to_numeric(rt_raw[qty_col], errors='coerce').fillna(1).astype(int)
        rt_raw['_clean_part'] = rt_raw[part_col].apply(self.clean_part)
        rt_agg = rt_raw.groupby('_clean_part')[qty_col].sum().reset_index()
        rt_agg.columns = ['Part', 'RT_Qty']
        rt_agg = rt_agg[rt_agg['Part'] != '']
        # Match RT scans against detail rows
        all_detail['_matched'] = False
        all_detail['_match_status'] = ''
        new_unmatched = []
        for _, rt_row in rt_agg.iterrows():
            rt_part = rt_row['Part']
            rt_qty = int(rt_row['RT_Qty'])
            mask_iet = (all_detail['_clean_iet'] == rt_part) & (~all_detail['_matched'])
            mask_pn = (all_detail['_clean_pn'] == rt_part) & (~all_detail['_matched'])
            matching_idx = all_detail[mask_iet].index.tolist()
            if not matching_idx:
                matching_idx = all_detail[mask_pn].index.tolist()
            if not matching_idx:
                new_unmatched.append({'Part': rt_part + '[', 'Qty': rt_qty})
                continue
            simple_qty = all_detail.loc[matching_idx, 'return_qty'].sum()
            simple_qty = int(simple_qty) if pd.notna(simple_qty) else len(matching_idx)
            if rt_qty >= simple_qty:
                all_detail.loc[matching_idx, '_matched'] = True
                all_detail.loc[matching_idx, '_match_status'] = 'full'
                excess = rt_qty - simple_qty
                if excess > 0:
                    new_unmatched.append({'Part': rt_part + '[', 'Qty': excess})
            else:
                claimed = 0
                for idx in matching_idx:
                    if claimed >= rt_qty:
                        break
                    row_qty = all_detail.at[idx, 'return_qty']
                    row_qty = int(row_qty) if pd.notna(row_qty) else 1
                    all_detail.at[idx, '_matched'] = True
                    all_detail.at[idx, '_match_status'] = 'partial'
                    claimed += row_qty
        # Split into tabs
        prev_received_df = all_detail[all_detail['_match_status'] == 'full'].copy()
        ready_df = all_detail[all_detail['_match_status'] == 'partial'].copy()
        remaining_df = all_detail[~all_detail['_matched']].copy()
        drop_cols = ['_clean_iet', '_clean_pn', '_matched', '_match_status']
        prev_received_df = prev_received_df.drop(columns=drop_cols, errors='ignore')
        ready_df = ready_df.drop(columns=drop_cols, errors='ignore')
        remaining_df = remaining_df.drop(columns=drop_cols, errors='ignore')
        # Build unmatched tab - carry forward existing + add new
        new_unmatched_df = pd.DataFrame(new_unmatched)
        if len(existing_unmatched) > 0 and len(new_unmatched_df) > 0:
            for col in existing_unmatched.columns:
                if col not in new_unmatched_df.columns:
                    new_unmatched_df[col] = ''
            new_unmatched_df = new_unmatched_df[existing_unmatched.columns]
            unmatched_df = pd.concat([existing_unmatched, new_unmatched_df], ignore_index=True)
            unmatched_df = unmatched_df.drop_duplicates(subset=['Part'], keep='first')
        elif len(existing_unmatched) > 0:
            unmatched_df = existing_unmatched
        elif len(new_unmatched_df) > 0:
            unmatched_df = new_unmatched_df
        else:
            unmatched_df = pd.DataFrame(columns=['Part', 'Qty'])
        total_out = len(remaining_df) + len(ready_df) + len(prev_received_df)
        stats = {'total_in': total_in, 'remaining': len(remaining_df), 'ready': len(ready_df), 'prev_received': len(prev_received_df), 'unmatched': len(unmatched_df), 'total_out': total_out}
        output_file = os.path.join(os.path.dirname(simple_file), "Reconciled_{}.xlsx".format(datetime.now().strftime('%Y%m%d_%H%M%S')))
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            remaining_df.to_excel(writer, sheet_name='IE Tire', index=False)
            ready_df.to_excel(writer, sheet_name='Ready to Receive', index=False)
            unmatched_df.to_excel(writer, sheet_name='Unmatched', index=False)
            prev_received_df.to_excel(writer, sheet_name='Previously Received', index=False)
        self.format_workbook(output_file)
        return output_file, stats

    def format_workbook(self, file_path):
        wb = load_workbook(file_path)
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font = Font(color='FFFFFF', bold=True)
        red = PatternFill('solid', fgColor='FFC7CE')
        yellow = PatternFill('solid', fgColor='FFEB9C')
        green = PatternFill('solid', fgColor='C6EFCE')
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            if ws.title == 'Ready to Receive':
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = yellow
            elif ws.title == 'Unmatched':
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = red
            elif ws.title == 'Previously Received':
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = green
        wb.save(file_path)

if __name__ == '__main__':
    root = tk.Tk()
    ReconcilerApp(root)
    root.mainloop()

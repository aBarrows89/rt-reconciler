import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
import os
import threading

class ReconcilerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("RT Reconciler")
        self.root.geometry("520x400")
        self.root.resizable(False, False)
        self.root.configure(bg='#f0f0f0')
        self.simple_file = tk.StringVar()
        self.rt_file = tk.StringVar()
        self.create_widgets()
    
    def create_widgets(self):
        # Title
        title = tk.Label(self.root, text="RT vs Simple Reconciler", font=("Arial", 18, "bold"), bg='#f0f0f0')
        title.pack(pady=20)
        
        # Simple Workbook
        frame1 = tk.Frame(self.root, bg='#f0f0f0', pady=10)
        frame1.pack(fill="x", padx=20)
        tk.Label(frame1, text="Simple Workbook:", font=("Arial", 10), bg='#f0f0f0').pack(anchor="w")
        f1 = tk.Frame(frame1, bg='#f0f0f0')
        f1.pack(fill="x", pady=5)
        tk.Entry(f1, textvariable=self.simple_file, width=55, font=("Arial", 9)).pack(side="left", fill="x", expand=True)
        tk.Button(f1, text="Browse...", command=self.browse_simple, width=10).pack(side="right", padx=5)
        
        # RT Export
        frame2 = tk.Frame(self.root, bg='#f0f0f0', pady=10)
        frame2.pack(fill="x", padx=20)
        tk.Label(frame2, text="RT Comparison:", font=("Arial", 10), bg='#f0f0f0').pack(anchor="w")
        f2 = tk.Frame(frame2, bg='#f0f0f0')
        f2.pack(fill="x", pady=5)
        tk.Entry(f2, textvariable=self.rt_file, width=55, font=("Arial", 9)).pack(side="left", fill="x", expand=True)
        tk.Button(f2, text="Browse...", command=self.browse_rt, width=10).pack(side="right", padx=5)
        
        # Progress
        self.progress = ttk.Progressbar(self.root, mode="indeterminate", length=450)
        self.progress.pack(pady=20)
        
        # Status
        self.status_var = tk.StringVar(value="Select files and click Reconcile")
        tk.Label(self.root, textvariable=self.status_var, font=("Arial", 10), bg='#f0f0f0').pack(pady=5)
        
        # Big visible button
        self.btn = tk.Button(
            self.root, 
            text="RECONCILE", 
            font=("Arial", 14, "bold"),
            width=20,
            height=2,
            bg='#4472C4',
            fg='white',
            activebackground='#3461b3',
            activeforeground='white',
            cursor='hand2',
            command=self.start_reconcile
        )
        self.btn.pack(pady=20)
    
    def browse_simple(self):
        f = filedialog.askopenfilename(title="Select Simple Workbook", filetypes=[("Excel", "*.xlsx *.xls")])
        if f: self.simple_file.set(f)
    
    def browse_rt(self):
        f = filedialog.askopenfilename(title="Select RT Comparison", filetypes=[("Excel", "*.xlsx *.xls")])
        if f: self.rt_file.set(f)
    
    def start_reconcile(self):
        if not self.simple_file.get() or not self.rt_file.get():
            messagebox.showerror("Error", "Please select both files")
            return
        self.btn.config(state='disabled')
        self.status_var.set("Processing...")
        self.root.update()
        try:
            output, stats = self.reconcile(self.simple_file.get(), self.rt_file.get())
            self.on_complete(output, stats)
        except Exception as e:
            import traceback
            self.on_error(f"{str(e)}\n\n{traceback.format_exc()}")
    
    def on_complete(self, output_file, stats):
        self.btn.config(state='normal')
        self.status_var.set("Complete!")
        
        msg = f"""Reconciliation complete!

SIMPLE Total: {stats['simple']}
RT Total: {stats['rt']}
Variance: {stats['variance']}

Output: {os.path.basename(output_file)}

Tabs: IE Tire | Ready to Receive | Unmatched | Previously Received

Open now?"""
        
        if messagebox.askyesno("Success", msg):
            os.startfile(output_file)
    
    def on_error(self, msg):
        self.btn.config(state='normal')
        self.status_var.set("Error")
        messagebox.showerror("Error", msg)
    
    def reconcile(self, simple_file, rt_file):
        # Load detail data from Simple Workbook
        detail_df = pd.read_excel(simple_file, sheet_name='IE Tire')
        
        # Load RT Comparison - this has the correct SIMPLE, RT, and DIFF already
        rt_df = pd.read_excel(rt_file, sheet_name=0)
        
        # Standardize column names - handle 3 or 4 column files
        if len(rt_df.columns) == 4:
            rt_df.columns = ['IET #', 'SIMPLE', 'RT', 'DIFF']
        elif len(rt_df.columns) == 3:
            rt_df.columns = ['IET #', 'SIMPLE', 'RT']
            rt_df['SIMPLE'] = pd.to_numeric(rt_df['SIMPLE'], errors='coerce').fillna(0).astype(int)
            rt_df['RT'] = pd.to_numeric(rt_df['RT'], errors='coerce').fillna(0).astype(int)
            rt_df['DIFF'] = rt_df['SIMPLE'] - rt_df['RT']
        else:
            raise ValueError(f"RT file has {len(rt_df.columns)} columns, expected 3 or 4")
        rt_df = rt_df.dropna(subset=['IET #'])
        
        # Use RT file's numbers directly
        stats = {
            'simple': int(rt_df['SIMPLE'].sum()),
            'rt': int(rt_df['RT'].sum()),
            'variance': int(rt_df['DIFF'].sum())
        }
        
        # Build diff dict from RT file (only positive DIFF = Simple has more than RT)
        diff_dict = {}
        for _, row in rt_df.iterrows():
            diff = int(row['DIFF']) if pd.notna(row['DIFF']) else 0
            if diff > 0:
                diff_dict[self.normalize_sku(row['IET #'])] = diff
        
        # Add Variance Qty to detail
        detail_df = self.add_variance_qty(detail_df, diff_dict)
        
        output_file = os.path.join(os.path.dirname(simple_file), f"Reconciled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        # Split RT data into filtered views
        # Ready to Receive: exists in both but quantities don't match (DIFF > 0, RT > 0, SIMPLE > 0)
        ready_df = rt_df[(rt_df['DIFF'] > 0) & (rt_df['RT'] > 0) & (rt_df['SIMPLE'] > 0)].copy()
        # Unmatched: in RT but not in Simple (SIMPLE = 0, RT > 0)
        unmatched_df = rt_df[(rt_df['SIMPLE'] == 0) & (rt_df['RT'] > 0)].copy()
        # Previously Received: fully matched (DIFF = 0)
        received_df = rt_df[rt_df['DIFF'] == 0].copy()

        # Output tabs
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # IE Tire - full detail with variance
            detail_df.to_excel(writer, sheet_name='IE Tire', index=False)
            
            # Ready to Receive - items still waiting (DIFF > 0)
            ready_df.to_excel(writer, sheet_name='Ready to Receive', index=False)
            
            # Unmatched - items with variances (DIFF > 0)
            unmatched_df.to_excel(writer, sheet_name='Unmatched', index=False)
            
            # Previously Received - matched items (DIFF = 0)
            received_df.to_excel(writer, sheet_name='Previously Received', index=False)
        
        self.format_workbook(output_file)
        
        return output_file, stats
    
    @staticmethod
    def normalize_sku(val):
        """Normalize SKU values to prevent mismatches from whitespace, case, or float casting."""
        s = str(val).strip().upper()
        # Kill trailing .0 from pandas float conversion
        if s.endswith('.0'):
            s = s[:-2]
        return s

    def add_variance_qty(self, detail_df, diff_dict):
        detail_df = detail_df.copy()
        detail_df['Variance Qty'] = 0
        remaining = diff_dict.copy()
        
        for idx, row in detail_df.iterrows():
            iet = self.normalize_sku(row['IET #'])
            if iet in remaining and remaining[iet] > 0:
                row_qty = int(row['return_qty']) if pd.notna(row['return_qty']) else 1
                assign = min(row_qty, remaining[iet])
                detail_df.at[idx, 'Variance Qty'] = assign
                remaining[iet] -= assign
        
        return detail_df
    
    def format_workbook(self, file_path):
        wb = load_workbook(file_path)
        
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font = Font(color='FFFFFF', bold=True)
        red = PatternFill('solid', fgColor='FFC7CE')
        
        for ws in wb.worksheets:
            # Headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
            
            # Highlight Variance Qty > 0 in IE Tire
            if ws.title == 'IE Tire':
                var_col = None
                for i, cell in enumerate(ws[1], 1):
                    if cell.value == 'Variance Qty':
                        var_col = i
                        break
                
                if var_col:
                    for row in ws.iter_rows(min_row=2):
                        val = row[var_col-1].value
                        if val and val > 0:
                            for cell in row:
                                cell.fill = red
            
            # Highlight Ready to Receive rows yellow (qty mismatch)
            if ws.title == 'Ready to Receive':
                yellow = PatternFill('solid', fgColor='FFEB9C')
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = yellow
            
            # Highlight Unmatched rows red (not in RT at all)
            if ws.title == 'Unmatched':
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = red
            
            # Highlight all data rows in Previously Received green
            if ws.title == 'Previously Received':
                green = PatternFill('solid', fgColor='C6EFCE')
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.fill = green
        
        wb.save(file_path)

if __name__ == '__main__':
    root = tk.Tk()
    ReconcilerApp(root)
    root.mainloop()
